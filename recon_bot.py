# =============================================================
# recon_bot.py  —  Main reconciliation bot
# =============================================================

import os
import shutil
import logging
from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from thefuzz import fuzz

from config import (INTAKE_FOLDER, PROCESSED_FOLDER, OUTPUT_FOLDER,
                    MEDICAL_AID_MAP, FUZZY_THRESHOLD)
from parsers import parse_remittance, load_excel_claims
from reason_engine import lookup_reason

# ── Logging ───────────────────────────────────────────────────
os.makedirs(OUTPUT_FOLDER,    exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

log_path = os.path.join(OUTPUT_FOLDER,
    f"recon_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
GREEN_FILL   = PatternFill("solid", fgColor="E2EFDA")
AMBER_FILL   = PatternFill("solid", fgColor="FFF2CC")
RED_FILL     = PatternFill("solid", fgColor="FCE4D6")
BLUE_FILL    = PatternFill("solid", fgColor="DEEAF1")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

COLOUR_MAP = {
    "GREEN": GREEN_FILL,
    "AMBER": AMBER_FILL,
    "RED"  : RED_FILL,
}

def _style_header(ws, row_num, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = THIN_BORDER

def _auto_width(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ── Matching ──────────────────────────────────────────────────

def _norm(s):
    return str(s).strip().upper().replace("  ", " ")

def _match_by_invoice(pdf_tx, excel_claims):
    inv = _norm(pdf_tx.get("invoice_num", ""))
    if not inv or inv in ("", "NAN"):
        return None, None
    for claim in excel_claims:
        if not claim["matched"] and _norm(claim["claim_ref"]) == inv:
            return claim, "invoice"
    return None, None

def _match_by_member_number(pdf_tx, excel_claims):
    """Match on medical aid membership number."""
    pdf_num = _norm(pdf_tx.get("member_id", "") or pdf_tx.get("member_num", ""))
    if not pdf_num or pdf_num in ("", "NAN"):
        return None, None
    pdf_date = pdf_tx.get("treat_date")
    for claim in excel_claims:
        if claim["matched"]:
            continue
        excel_num = _norm(claim.get("member_number", ""))
        if excel_num and excel_num == pdf_num:
            if pdf_date and claim["visit_date"] == pdf_date:
                return claim, "member_number"
    return None, None

def _match_by_name_date(pdf_tx, excel_claims):
    pdf_name = _norm(pdf_tx.get("patient_name", "")
                     or pdf_tx.get("member_name", ""))
    pdf_date = pdf_tx.get("treat_date")
    if not pdf_name or not pdf_date:
        return None, None
    best_score = 0
    best_claim = None
    for claim in excel_claims:
        if claim["matched"]:
            continue
        if claim["visit_date"] != pdf_date:
            continue
        score = fuzz.token_sort_ratio(pdf_name, _norm(claim["member_name"]))
        if score > best_score:
            best_score = score
            best_claim = claim
    if best_score >= FUZZY_THRESHOLD:
        return best_claim, "fuzzy_name"
    return None, None


# ── Output builder ────────────────────────────────────────────

def build_output(run_date, pdf_transactions, excel_claims,
                 match_log, error_log):

    wb = Workbook()
    wb.remove(wb.active)

    # ── Tab 1: Summary ────────────────────────────────────────
    ws_sum = wb.create_sheet("1 - Summary")

    total_shortfall = round(sum(
        t["shortfall_amt"] for t in pdf_transactions), 2)
    n_reconciled = sum(
        1 for t in pdf_transactions if t["shortfall_amt"] == 0 and t["matched"])
    n_shortfall = sum(
        1 for t in pdf_transactions if t["shortfall_amt"] > 0 and t["matched"])
    n_unmatched_pdf = sum(1 for t in pdf_transactions if not t["matched"])
    n_unmatched_xl  = sum(1 for c in excel_claims    if not c["matched"])

    by_aid = {}
    for t in pdf_transactions:
        aid = t.get("medical_aid", "Unknown")
        if aid not in by_aid:
            by_aid[aid] = {"transactions": 0, "shortfall": 0.0}
        by_aid[aid]["transactions"] += 1
        by_aid[aid]["shortfall"]    += t.get("shortfall_amt", 0)

    summary_rows = [
        ["Medical Aid Reconciliation — Run Summary", ""],
        ["Run date", run_date.strftime("%d %B %Y %H:%M")],
        [""],
        ["TOTALS", ""],
        ["PDF transactions parsed",   len(pdf_transactions)],
        ["Excel claims loaded",        len(excel_claims)],
        ["Matched (invoice)",          sum(1 for m in match_log if m["match_method"] == "invoice")],
        ["Matched (member number)",    sum(1 for m in match_log if m["match_method"] == "member_number")],
        ["Matched (name + date)",      sum(1 for m in match_log if m["match_method"] == "fuzzy_name")],
        ["Unmatched PDF lines",        n_unmatched_pdf],
        ["Unmatched Excel claims",     n_unmatched_xl],
        [""],
        ["SHORTFALL SUMMARY", ""],
        ["Total shortfall value (USD)", total_shortfall],
        ["Fully reconciled (no shortfall)", n_reconciled],
        ["Shortfalls requiring action",     n_shortfall],
        [""],
        ["BY MEDICAL AID", "Transactions | Shortfall (USD)"],
    ]
    for aid, stats in sorted(by_aid.items()):
        summary_rows.append([aid,
            f"{stats['transactions']} transactions  |  "
            f"USD {round(stats['shortfall'], 2):.2f} shortfall"])
    summary_rows += [[""], ["Errors encountered", len(error_log)]]

    for r, row in enumerate(summary_rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=14, color="1F4E79")
            elif val in ("TOTALS", "SHORTFALL SUMMARY", "BY MEDICAL AID"):
                cell.font = Font(bold=True, size=11)
                cell.fill = BLUE_FILL
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 38

    # ── Tab 2: Fully Reconciled ───────────────────────────────
    ws_rec = wb.create_sheet("2 - Fully Reconciled")
    rec_h = ["Medical Aid", "Patient Name", "Member Name", "Member Number",
             "Treatment Date", "Invoice Number", "Tariff Code",
             "Claimed (USD)", "Accepted (USD)", "Shortfall (USD)",
             "Pay To You", "Match Method"]
    ws_rec.append(rec_h)
    _style_header(ws_rec, 1, len(rec_h))
    for tx in pdf_transactions:
        if tx["shortfall_amt"] == 0 and tx["matched"]:
            r = ws_rec.max_row + 1
            ws_rec.append([
                tx.get("medical_aid", ""),
                tx.get("patient_name", ""),
                tx.get("member_name", ""),
                tx.get("member_id", "") or tx.get("member_num", ""),
                str(tx.get("treat_date", "")),
                tx.get("invoice_num", ""),
                tx.get("tariff_code", ""),
                tx.get("claimed_amt", 0),
                tx.get("accepted_amt", 0),
                0,
                tx.get("pay_to_you", 0),
                tx.get("match_method", ""),
            ])
            for c in range(1, len(rec_h) + 1):
                ws_rec.cell(r, c).fill   = GREEN_FILL
                ws_rec.cell(r, c).border = THIN_BORDER
    _auto_width(ws_rec)

    # ── Tab 3: Shortfalls for Action ─────────────────────────
    ws_sf = wb.create_sheet("3 - Shortfalls for Action")
    sf_h = ["Medical Aid", "Patient Name", "Member Name", "Member Number",
            "Treatment Date", "Invoice Number", "Tariff Code",
            "Claimed (USD)", "Accepted (USD)", "Shortfall (USD)",
            "Reason Code", "Reason Description",
            "Classification", "Action Required",
            "Patient Phone", "Match Method"]
    ws_sf.append(sf_h)
    _style_header(ws_sf, 1, len(sf_h))
    for tx in pdf_transactions:
        if tx["shortfall_amt"] > 0 and tx["matched"]:
            desc, cls, action, colour = lookup_reason(
                tx.get("reason_code", ""),
                tx.get("medical_aid", ""),
                tx.get("reason_desc", ""))
            fill = COLOUR_MAP.get(colour, AMBER_FILL)
            r = ws_sf.max_row + 1
            ws_sf.append([
                tx.get("medical_aid", ""),
                tx.get("patient_name", ""),
                tx.get("member_name", ""),
                tx.get("member_id", "") or tx.get("member_num", ""),
                str(tx.get("treat_date", "")),
                tx.get("invoice_num", ""),
                tx.get("tariff_code", ""),
                tx.get("claimed_amt", 0),
                tx.get("accepted_amt", 0),
                tx.get("shortfall_amt", 0),
                tx.get("reason_code", ""),
                desc,
                cls,
                action,
                tx.get("phone", ""),
                tx.get("match_method", ""),
            ])
            for c in range(1, len(sf_h) + 1):
                ws_sf.cell(r, c).fill   = fill
                ws_sf.cell(r, c).border = THIN_BORDER
    _auto_width(ws_sf)

    # ── Tab 4: Unmatched Records ──────────────────────────────
    ws_um = wb.create_sheet("4 - Unmatched Records")
    um_h = ["Source", "Medical Aid", "Patient / Member Name",
            "Member Number", "Treatment Date", "Invoice Number",
            "Claimed (USD)", "Shortfall (USD)",
            "Possible Reason", "Notes"]
    ws_um.append(um_h)
    _style_header(ws_um, 1, len(um_h))
    for tx in pdf_transactions:
        if not tx["matched"]:
            r = ws_um.max_row + 1
            ws_um.append([
                "PDF remittance",
                tx.get("medical_aid", ""),
                tx.get("patient_name", "") or tx.get("member_name", ""),
                tx.get("member_id", "") or tx.get("member_num", ""),
                str(tx.get("treat_date", "")),
                tx.get("invoice_num", ""),
                tx.get("claimed_amt", 0),
                tx.get("shortfall_amt", 0),
                "Claim not found in Client Data.xlsx",
                "Check if patient was recorded or claim was submitted",
            ])
            for c in range(1, len(um_h) + 1):
                ws_um.cell(r, c).fill   = AMBER_FILL
                ws_um.cell(r, c).border = THIN_BORDER
    for claim in excel_claims:
        if not claim["matched"]:
            r = ws_um.max_row + 1
            ws_um.append([
                "Excel claim record",
                claim.get("payer", ""),
                claim.get("member_name", ""),
                claim.get("member_number", ""),
                str(claim.get("visit_date", "")),
                claim.get("claim_ref", ""),
                claim.get("claimed_usd", 0),
                "",
                "No remittance line found for this claim",
                "Claim may not yet be processed, or was not submitted",
            ])
            for c in range(1, len(um_h) + 1):
                ws_um.cell(r, c).fill   = RED_FILL
                ws_um.cell(r, c).border = THIN_BORDER
    _auto_width(ws_um)

    # ── Tab 5: Error Log ──────────────────────────────────────
    ws_err = wb.create_sheet("5 - Error Log")
    err_h = ["Timestamp", "File", "Error"]
    ws_err.append(err_h)
    _style_header(ws_err, 1, len(err_h))
    for entry in error_log:
        ws_err.append(entry)
    _auto_width(ws_err)

    # ── Tab 6: Action Tracker ─────────────────────────────────
    ws_at = wb.create_sheet("6 - Action Tracker")
    at_h = ["Status", "Medical Aid", "Patient Name", "Member Name",
            "Member Number", "Treatment Date", "Invoice Number",
            "Shortfall (USD)", "Reason Code", "Reason Description",
            "Classification", "Action Required",
            "Patient Phone", "Date Actioned", "Notes"]
    ws_at.append(at_h)
    _style_header(ws_at, 1, len(at_h))
    for tx in pdf_transactions:
        if tx["shortfall_amt"] > 0 and tx["matched"]:
            desc, cls, action, colour = lookup_reason(
                tx.get("reason_code", ""),
                tx.get("medical_aid", ""),
                tx.get("reason_desc", ""))
            r = ws_at.max_row + 1
            ws_at.append([
                "Open",
                tx.get("medical_aid", ""),
                tx.get("patient_name", ""),
                tx.get("member_name", ""),
                tx.get("member_id", "") or tx.get("member_num", ""),
                str(tx.get("treat_date", "")),
                tx.get("invoice_num", ""),
                tx.get("shortfall_amt", 0),
                tx.get("reason_code", ""),
                desc,
                cls,
                action,
                tx.get("phone", ""),
                "",
                "",
            ])
            for c in range(1, len(at_h) + 1):
                cell = ws_at.cell(r, c)
                cell.border = THIN_BORDER
                if c == 1:
                    cell.fill = AMBER_FILL
                    cell.font = Font(bold=True)
    ws_at.freeze_panes = "A2"
    _auto_width(ws_at)

    return wb


# ── Main ──────────────────────────────────────────────────────

def run():
    run_start = datetime.now()
    error_log = []
    match_log = []

    log.info("=" * 60)
    log.info("Medical Aid Reconciliation Bot — START")
    log.info("=" * 60)

    log.info("Loading Client Data.xlsx ...")
    try:
        excel_claims = load_excel_claims()
        log.info(f"  Loaded {len(excel_claims)} medical aid claim records")
    except Exception as e:
        log.error(f"FATAL: Cannot load Client Data.xlsx — {e}")
        return

    pdf_files = [f for f in os.listdir(INTAKE_FOLDER)
                 if f.lower().endswith(".pdf")]
    log.info(f"Found {len(pdf_files)} PDF file(s)")

    all_pdf_transactions = []
    processed_files      = []

    for filename in pdf_files:
        pdf_path = os.path.join(INTAKE_FOLDER, filename)
        log.info(f"  Processing: {filename}")
        medical_aid_name = "Unknown"
        for key, name in MEDICAL_AID_MAP.items():
            if key in filename.lower():
                medical_aid_name = name
                break
        try:
            transactions = parse_remittance(pdf_path, medical_aid_name)
            log.info(f"    Parsed {len(transactions)} transaction lines")
            all_pdf_transactions.extend(transactions)
            processed_files.append((filename, pdf_path))
        except Exception as e:
            log.error(f"    Error parsing {filename}: {e}")
            error_log.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                filename, str(e)])

    log.info("Matching transactions ...")
    for tx in all_pdf_transactions:
        claim, method = _match_by_invoice(tx, excel_claims)
        if claim is None:
            claim, method = _match_by_member_number(tx, excel_claims)
        if claim is None:
            claim, method = _match_by_name_date(tx, excel_claims)

        if claim is not None:
            tx["matched"]      = True
            tx["match_method"] = method
            tx["phone"]        = claim.get("phone", "")
            claim["matched"]   = True
            match_log.append({"match_method": method,
                              "invoice": tx.get("invoice_num", ""),
                              "patient": tx.get("patient_name", "")})
            log.info(f"    MATCH ({method}): {tx.get('invoice_num','')} — {tx.get('patient_name','')}")
        else:
            tx["matched"]      = False
            tx["match_method"] = "unmatched"
            log.warning(f"    NO MATCH: {tx.get('invoice_num','')} — {tx.get('patient_name','')} on {tx.get('treat_date','')}")

    log.info("Building output workbook ...")
    wb = build_output(run_start, all_pdf_transactions,
                      excel_claims, match_log, error_log)

    out_name = f"RECON_{run_start.strftime('%Y%m%d_%H%M%S')}_ALL.xlsx"
    out_path = os.path.join(OUTPUT_FOLDER, out_name)
    wb.save(out_path)
    log.info(f"  Output saved: {out_path}")

    dated_archive = os.path.join(
        PROCESSED_FOLDER, run_start.strftime("%Y%m%d"))
    os.makedirs(dated_archive, exist_ok=True)
    for filename, pdf_path in processed_files:
        shutil.move(pdf_path, os.path.join(dated_archive, filename))
        log.info(f"  Archived: {filename}")

    log.info("=" * 60)
    log.info("RECONCILIATION COMPLETE")
    log.info(f"  PDF transactions : {len(all_pdf_transactions)}")
    log.info(f"  Excel claims     : {len(excel_claims)}")
    log.info(f"  Matched          : {len(match_log)}")
    log.info(f"  Unmatched PDF    : {sum(1 for t in all_pdf_transactions if not t['matched'])}")
    log.info(f"  Unmatched Excel  : {sum(1 for c in excel_claims if not c['matched'])}")
    log.info(f"  Errors           : {len(error_log)}")
    log.info(f"  Output           : {out_name}")
    log.info("=" * 60)


if __name__ == "__main__":
    run()