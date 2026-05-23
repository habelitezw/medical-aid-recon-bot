# =============================================================
# create_test_data.py
# Run once to generate test Excel + test PDF remittances
# covering ALL output tab scenarios.
# Usage: python create_test_data.py
# =============================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

OUT_DIR = r"D:\Medical Aid\TestData"
os.makedirs(OUT_DIR, exist_ok=True)
styles = getSampleStyleSheet()

# ── Test scenario design ───────────────────────────────────────────────────────
#
# Tab 2 Fully Reconciled    : Peter Ncube (FMH, invoice match)
#                             Clara Osei  (Bonvie, member number match)
#                             John Moyo   (Cimas, invoice match)
#                             Mary Dube   (Cimas, invoice match)
#
# Tab 3 Shortfalls for Action:
#   Grace Mutasa  — Cimas,   tariff difference        (reason 6)
#   Tafara Choto  — FMH,     benefit exhausted        (reason D)
#   Ruth Banda    — Alliance, not a covered benefit   (reason E)
#   David Sithole — Cimas,   data/submission error    (reason 40)
#
# Tab 4 Unmatched PDF  : Unregistered Patient (Cimas)
#                        Unknown Member       (FMH)
# Tab 4 Unmatched Excel: Agnes Mokoena (CIMAS — no remittance sent)
#                        Brian Mwale   (FMH   — no remittance sent)
#
# Tab 5 Error Log      : corrupt_test.pdf (unreadable file)
#
# Tab 6 Action Tracker : auto-populated from Tab 3 shortfalls

EXCEL_CLAIMS = [
    # (date, payer, member_name, member_number, claim_ref, claimed, phone)
    (date(2026,3,1),  "CIMAS",   "John Moyo",     "CIMAS-001", "0931-99001-11111-AA1", 50.00,  "+263771000001"),
    (date(2026,3,5),  "CIMAS",   "Mary Dube",     "CIMAS-002", "0931-99002-22222-BB2", 70.00,  "+263771000002"),
    (date(2026,3,10), "FMH",     "Peter Ncube",   "FMH-001",   "0931-99003-33333-CC3", 80.00,  "+263771000003"),
    (date(2026,3,12), "CIMAS",   "Grace Mutasa",  "CIMAS-003", "0931-99004-44444-DD4", 60.00,  "+263771000004"),
    (date(2026,3,15), "FMH",     "Tafara Choto",  "FMH-002",   "0931-99005-55555-EE5", 50.00,  "+263771000005"),
    (date(2026,3,18), "ALLIANCE","Ruth Banda",    "ALI-001",   "ALI-99006-66666-FF6",  80.00,  "+263771000006"),
    (date(2026,3,20), "CIMAS",   "David Sithole", "CIMAS-004", "0931-99007-77777-GG7", 50.00,  "+263771000007"),
    # No remittance — will appear as unmatched Excel records
    (date(2026,3,22), "CIMAS",   "Agnes Mokoena", "CIMAS-005", "0931-99008-88888-HH8", 70.00,  "+263771000008"),
    (date(2026,3,25), "FMH",     "Brian Mwale",   "FMH-003",   "0931-99009-99999-II9", 50.00,  "+263771000009"),
    # Member number match — no invoice in Excel
    (date(2026,3,28), "BONVIE",  "Clara Osei",    "BON-001",   "",                     60.00,  "+263771000010"),
]


# ── Build Excel ────────────────────────────────────────────────────────────────

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Data"

    headers = [
        "Date", "Col2", "Col3", "Col4", "Col5", "Col6",
        "Surname", "First Name", "Col9", "Col10", "Col11",
        "Payer", "Medical Aid Number", "Col14",
        "Primary Medical Aid Member",
        "Col16","Col17","Col18","Col19","Col20",
        "Col21","Col22","Col23","Col24","Col25",
        "Col26","Col27",
        "NH263 Claim reference",
        "MEDICAL AID Amount claimed USD",
        "Col30","Col31",
        "Medical Aid Amount remitted USD",
        "Medical Aid Discrepancy USD",
        "Col34","Col35",
        "Medical Aid          Reason for discrepancy",
        "Claim Status",
        "Patient ID ",
        "Phone Number",
    ]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont

    for dt, payer, member, mem_num, ref, claimed, phone in EXCEL_CLAIMS:
        row = [""] * len(headers)
        row[0]  = dt
        row[11] = payer
        row[12] = mem_num
        row[14] = member
        row[27] = ref
        row[28] = claimed
        row[38] = phone
        ws.append(row)

    out = os.path.join(OUT_DIR, "Test_Client_Data.xlsx")
    wb.save(out)
    print(f"  Created: {out}")


# ── PDF builder ────────────────────────────────────────────────────────────────
# We use the Alliance table-based format for ALL test PDFs because:
#   - Alliance parser reads clean tables reliably
#   - Avoids the FMH text-wrap complexity in synthetic PDFs
#   - All aids will be named correctly from the filename

def _make_pdf(filename, aid_name, payment_ref, transactions, reason_codes):
    """
    Build a table-based remittance PDF in Alliance-style format.
    transactions: list of dicts:
      member_name, member_id, patient_name, treat_date,
      invoice, code, claimed, accepted, shortfall, reason_code
    reason_codes: dict {code: description}
    """
    path = os.path.join(OUT_DIR, filename)
    doc  = SimpleDocTemplate(path, pagesize=A4,
                             leftMargin=10*mm, rightMargin=10*mm,
                             topMargin=12*mm, bottomMargin=12*mm)
    elems = []

    elems.append(Paragraph(aid_name, styles["Heading1"]))
    elems.append(Paragraph("Claims Remittance Advice", styles["Heading2"]))
    elems.append(Paragraph(
        f"Payment Reference: {payment_ref}    Date of payment run: 31/03/2026",
        styles["Normal"]))
    elems.append(Paragraph(
        "Provider: WASHAYA JENNIFER (19097)",
        styles["Normal"]))
    elems.append(Spacer(1, 5*mm))

    # Main claims table — exact Alliance column structure
    # IMPORTANT: Claim Date must be first — parser uses it to identify data rows
    col_labels = ["Claim Date", "ClaimNo", "MemberId",
                  "Member Name", "Patient",
                  "Claimed", "Award\nAmount", "ShortFall\nAmount",
                  "ShortFall\nCode", "AFHOZ/\nDrugs",
                  "Diagnosis", "Invoice\nNumber", "EFTNumber"]
    col_widths  = [18*mm, 20*mm, 18*mm, 26*mm, 20*mm,
                   14*mm, 14*mm, 14*mm, 12*mm, 12*mm,
                   18*mm, 32*mm, 20*mm]

    table_data = [col_labels]
    for tx in transactions:
        table_data.append([
            tx["treat_date"],                                          # Claim Date
            f"CLM{tx['invoice'][-6:] if tx['invoice'] else '000000'}", # ClaimNo
            tx["member_id"],                                           # MemberId
            tx["member_name"],                                         # Member Name
            tx["patient_name"],                                        # Patient
            f"{tx['claimed']:.2f}",                                    # Claimed
            f"{tx['accepted']:.2f}",                                   # Award Amount
            f"{tx['shortfall']:.2f}",                                  # ShortFall Amount
            tx.get("reason_code", ""),                                 # ShortFall Code
            tx.get("code", ""),                                        # AFHOZ/Drugs
            "General",                                                 # Diagnosis
            tx["invoice"],                                             # Invoice Number
            f"EFT-{payment_ref[-4:]}",                                 # EFTNumber
        ])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 6.5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1),
         [colors.HexColor("#f7fbff"), colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("ALIGN",         (5,0), (7,-1),  "RIGHT"),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 5*mm))

    # Totals block
    total_claimed   = sum(tx["claimed"]   for tx in transactions)
    total_awarded   = sum(tx["accepted"]  for tx in transactions)
    total_shortfall = sum(tx["shortfall"] for tx in transactions)
    totals_data = [
        ["TotalAwarded :",   f"{total_awarded:.2f}",   ""],
        ["WithHolding Tax Total :", "0.00",             ""],
        ["EFT Total",        f"{total_claimed:.2f}",   ""],
    ]
    tt = Table(totals_data, colWidths=[60*mm, 30*mm, 100*mm])
    tt.setStyle(TableStyle([
        ("FONTSIZE",  (0,0), (-1,-1), 7),
        ("FONTNAME",  (0,0), (-1,0),  "Helvetica-Bold"),
        ("BACKGROUND",(0,0), (-1,-1), colors.HexColor("#DEEAF1")),
        ("GRID",      (0,0), (1,-1),  0.3, colors.grey),
    ]))
    elems.append(tt)
    elems.append(Spacer(1, 5*mm))

    # Reason codes table
    if reason_codes:
        rc_data = [["ShortFall Code", "ShortFall Description"]]
        for code, desc in reason_codes.items():
            rc_data.append([code, desc])
        rt = Table(rc_data, colWidths=[25*mm, 165*mm])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0),  colors.HexColor("#DEEAF1")),
            ("FONTNAME",   (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 7),
            ("GRID",       (0,0), (-1,-1), 0.3, colors.grey),
        ]))
        elems.append(rt)

    doc.build(elems)
    print(f"  Created: {path}")


def build_pdfs():

    # ── Cimas: 2 fully reconciled, 1 tariff shortfall,
    #           1 data error shortfall, 1 unmatched PDF line
    _make_pdf(
        filename    = "cimas_test_recon.pdf",
        aid_name    = "Cimas Medical Aid",
        payment_ref = "CIMAS-EFT-20260331",
        transactions=[
            dict(member_name="John Moyo",    member_id="CIMAS-001",
                 patient_name="John Moyo",   treat_date="01/03/2026",
                 invoice="0931-99001-11111-AA1", code="90030",
                 claimed=50.00, accepted=50.00, shortfall=0.00,
                 reason_code=""),
            dict(member_name="Mary Dube",    member_id="CIMAS-002",
                 patient_name="Mary Dube",   treat_date="05/03/2026",
                 invoice="0931-99002-22222-BB2", code="90030",
                 claimed=70.00, accepted=70.00, shortfall=0.00,
                 reason_code=""),
            dict(member_name="Grace Mutasa", member_id="CIMAS-003",
                 patient_name="Grace Mutasa", treat_date="12/03/2026",
                 invoice="0931-99004-44444-DD4", code="67228",
                 claimed=60.00, accepted=54.00, shortfall=6.00,
                 reason_code="6"),
            dict(member_name="David Sithole",member_id="CIMAS-004",
                 patient_name="David Sithole",treat_date="20/03/2026",
                 invoice="0931-99007-77777-GG7", code="90030",
                 claimed=50.00, accepted=0.00, shortfall=50.00,
                 reason_code="40"),
            # Unmatched PDF line — patient not in Excel
            dict(member_name="Unregistered Patient", member_id="CIMAS-999",
                 patient_name="Unregistered Patient", treat_date="25/03/2026",
                 invoice="0931-99099-99999-ZZ9", code="90030",
                 claimed=50.00, accepted=50.00, shortfall=0.00,
                 reason_code=""),
        ],
        reason_codes={
            "6":  "Amount claimed exceeds tariff amount",
            "40": "Duplicate claim — already processed",
        }
    )

    # ── FMH: 1 fully reconciled, 1 benefit exhausted, 1 unmatched PDF line
    _make_pdf(
        filename    = "fmh_test_recon.pdf",
        aid_name    = "First Mutual Health",
        payment_ref = "FMH-EFT-20260331",
        transactions=[
            dict(member_name="Peter Ncube",  member_id="FMH-001",
                 patient_name="Peter Ncube", treat_date="10/03/2026",
                 invoice="0931-99003-33333-CC3", code="90030",
                 claimed=80.00, accepted=80.00, shortfall=0.00,
                 reason_code=""),
            dict(member_name="Tafara Choto", member_id="FMH-002",
                 patient_name="Tafara Choto",treat_date="15/03/2026",
                 invoice="0931-99005-55555-EE5", code="92021",
                 claimed=50.00, accepted=40.00, shortfall=10.00,
                 reason_code="D"),
            # Unmatched PDF line
            dict(member_name="Unknown Member",member_id="FMH-999",
                 patient_name="Unknown Member",treat_date="29/03/2026",
                 invoice="0931-99099-88888-YY8", code="90030",
                 claimed=60.00, accepted=60.00, shortfall=0.00,
                 reason_code=""),
        ],
        reason_codes={"D": "Benefits exhausted"}
    )

    # ── Alliance: 1 shortfall (not a covered benefit)
    _make_pdf(
        filename    = "alliance_test_recon.pdf",
        aid_name    = "Alliance Health",
        payment_ref = "ALI-EFT-20260401",
        transactions=[
            dict(member_name="Ruth Banda",  member_id="ALI-001",
                 patient_name="Ruth Banda", treat_date="18/03/2026",
                 invoice="ALI-99006-66666-FF6", code="66826",
                 claimed=80.00, accepted=0.00, shortfall=80.00,
                 reason_code="E"),
        ],
        reason_codes={"E": "Not a covered benefit under this plan"}
    )

    # ── Bonvie: 1 fully reconciled (matched by member number — no invoice)
    _make_pdf(
        filename    = "bonvie_test_recon.pdf",
        aid_name    = "Bonvie",
        payment_ref = "BON-BATCH-20260331",
        transactions=[
            dict(member_name="Clara Osei",  member_id="BON-001",
                 patient_name="Clara Osei", treat_date="28/03/2026",
                 invoice="424999", code="66826",
                 claimed=60.00, accepted=60.00, shortfall=0.00,
                 reason_code=""),
        ],
        reason_codes={}
    )

    # ── Corrupt PDF — triggers Error Log entry ─────────────────────────────
    corrupt_path = os.path.join(OUT_DIR, "corrupt_test.pdf")
    with open(corrupt_path, "wb") as f:
        f.write(b"%PDF-1.4\n% This file is intentionally corrupt for testing\n"
                b"1 0 obj\n<< /Type /Catalog >>\nendobj\n"
                b"CORRUPTED DATA $$$$$ NOT A VALID PDF XREF\n")
    print(f"  Created: {corrupt_path}")


if __name__ == "__main__":
    print("Generating test data...")
    build_excel()
    build_pdfs()
    print(f"\nAll files created in: {OUT_DIR}")
    print("\nUpload to the web app:")
    print("  Box 1 (Excel): Test_Client_Data.xlsx")
    print("  Box 2 (PDFs) : cimas_test_recon.pdf, fmh_test_recon.pdf,")
    print("                 alliance_test_recon.pdf, bonvie_test_recon.pdf,")
    print("                 corrupt_test.pdf")
    print("\nExpected results:")
    print("  Tab 2 Fully Reconciled     : 4 rows (John Moyo, Mary Dube, Peter Ncube, Clara Osei)")
    print("  Tab 3 Shortfalls for Action: 4 rows (Grace, Tafara, Ruth, David)")
    print("  Tab 4 Unmatched Records    : 6 rows (2 PDF, 2 Excel, 2 unmatched)")
    print("  Tab 5 Error Log            : 1 row  (corrupt_test.pdf)")
    print("  Tab 6 Action Tracker       : 4 rows (open shortfalls)")